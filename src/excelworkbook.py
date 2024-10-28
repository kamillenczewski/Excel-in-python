from openpyxl import load_workbook, Workbook
from os.path import isfile as file_exists

from src.intervalutil import indices_range, limited_generator

class ExcelWorkbook:
    def __init__(self, path=None):
        self.path = path

        if file_exists(path):
            self.obj = load_workbook(path)
        else:
            self.obj = Workbook()

    def active_sheet(self):
        return self.obj.active
    
    def set_active_sheet(self, name):
        if name == None:
            return self.active_sheet()
        else:
            self.obj.active = self.obj[name]
            return self.active_sheet()
    
    def save(self):
        self.obj.save(self.path)

    def get(self, string_range=None, indices_gen=None, sheet_name=None):
        sheet = self.set_active_sheet(sheet_name)

        if string_range == None and indices_gen == None:
            raise ValueError('string_range and indices_gen can not simultaneously None!')

        if string_range != None:
            indices_gen = indices_range(string_range)

        for index in indices_gen:
            value = sheet[index].value

            if value == None:
                break

            yield sheet[index].value

    def get_item(self, index, sheet_name=None):
        return self.set_active_sheet(sheet_name)[index].value

    def set(self, values, string_range=None, indices_gen=None, sheet_name=None):
        sheet = self.set_active_sheet(sheet_name)

        if string_range == None and indices_gen == None:
            raise ValueError('string_range and indices_gen can not simultaneously None!')
        
        values_length = len(values)

        if string_range == None and indices_gen != None:
            indices = list(limited_generator(indices_gen, values_length))

        if string_range != None:
            indices_gen = indices_range(string_range)
            limited_gen = (index for _, index in zip(range(values_length), indices_gen))
            indices = list(limited_gen)

        for index, value in zip(indices, values):
            sheet[index] = value